"""
generate.py — Glorb CLI batch renderer + Freesound XLSX generator.

Usage examples:
    python generate.py                              # interactive mode selector
    python generate.py --mode glorb                 # single mode, default settings
    python generate.py --mode glorb retro insects   # multiple modes
    python generate.py --all                        # all 28 modes
    python generate.py --all --duration 60          # all modes, 60 s each
    python generate.py --all --quality studio --output-dir exports/glorb
    python generate.py --all --freesound            # also generate freesound XLSX
    python generate.py --mode glorb --energy 80 --brightness 30 --chaos 70

Options:
    --mode MODE [MODE ...]   Mode(s) to render (see list below)
    --all                    Render all modes
    --duration SECONDS       Duration in seconds (default: 30)
    --quality PRESET         standard | high (default) | studio | float
    --energy N               0–100 knob (default: 50)
    --brightness N           0–100 knob (default: 50)
    --chaos N                0–100 knob (default: 50)
    --output-dir DIR         Output folder (default: exports/glorb)
    --freesound              Generate Freesound bulk XLSX alongside audio
    --metadata-only          Regenerate XLSX without re-rendering audio
    --seed N                 Random seed for reproducible renders (default: random)
"""

import argparse
import random
import sys
import io
from pathlib import Path

import numpy as np
import soundfile as sf

sys.path.insert(0, str(Path(__file__).parent))
import main as bb

# ── Mode registry ────────────────────────────────────────────────────────────

MODES = {
    "glorb":     ("Glorb",      "Abstract electronic blips and bloops, the original Glorb sound"),
    "retro":     ("Retro",      "8-bit chiptune square and pulse waves"),
    "nature":    ("Nature",     "Rain, fire, insects, and forest textures"),
    "scifi":     ("Sci-Fi",     "Phasers, warp drives, and laser bursts"),
    "haptic":    ("Haptic",     "Tactile vibration pulses and clicks"),
    "radio":     ("Radio",      "AM/FM static, morse, and transmission artefacts"),
    "ui-pack":   ("UI Pack",    "Short UI notification sounds: clicks, pops, chimes"),
    "foley":     ("Foley",      "Footsteps, paper, keys, and everyday impacts"),
    "underwater":("Underwater", "Sonar pings, bubble trains, deep pressure tones"),
    "weather":   ("Weather",    "Thunder, lightning crackle, rain intensity layers"),
    "bell":      ("Bell",       "Marimba, tubular bells, and resonant metal strikes"),
    "bass":      ("Bass",       "Deep sub-bass pulses and 808-style low-end hits"),
    "glitch":    ("Glitch",     "Buffer corruption, bit-crush artefacts, digital stutter"),
    "pinball":   ("Pinball",    "Flippers, bumpers, drain, and mechanical clicks"),
    "horror":    ("Horror",     "Dissonant drones, breath, reversed tails, tension builders"),
    "granular":  ("Granular",   "Scattered grain clouds across noise and pitched sources"),
    "lofi":      ("Lo-Fi",      "Vinyl crackle, tape wow/flutter, degraded telephone bandwidth"),
    "modem":     ("Modem",      "Dial-up handshake tones, DTMF digits, FSK data bursts"),
    "insects":   ("Insects",    "Cricket chirps, cicada buzzing, grasshopper rasps"),
    "gamelan":   ("Gamelan",    "Balinese inharmonic metallophones with acoustic beating pairs"),
    "glass":     ("Glass",      "Brittle shards, singing rims, and crystalline resonances"),
    "clockwork": ("Clockwork",  "Ticks, ratchets, springs, and tiny mechanisms"),
    "creature":  ("Creature",   "Imaginary chirps, calls, growls, and breath"),
    "cat":       ("Cat",        "Feline meows, mewls, trills, purrs, and hisses"),
    "birds":     ("Birds",      "Melodic peeps, whistles, trills, warbles, and calls"),
    "electricity":("Electricity", "Arcs, transformer hum, relays, and rising charge"),
    "cave":      ("Cave",       "Drips, stones, subterranean wind, and deep rumbles"),
    "arp":       ("Arp",        "Synthesizer arpeggiator cycling minor/major/pentatonic chords"),
}

MODE_TAGS = {
    "glorb":      "electronic blip bloop generative synthesizer abstract digital beep tone",
    "retro":      "8-bit chiptune retro videogame square-wave chip pixel arcade nostalgic",
    "nature":     "nature rain fire insects forest organic procedural ambient",
    "scifi":      "sci-fi laser phaser spaceship futuristic electronic science-fiction beam",
    "haptic":     "haptic vibration click tactile pulse feedback notification ui interface",
    "radio":      "radio static transmission morse noise interference analog vintage shortwave",
    "ui-pack":    "ui interface notification click pop chime alert system feedback",
    "foley":      "foley sound-design footstep impact paper cloth everyday realistic",
    "underwater": "underwater sonar bubble ping ocean depth pressure aquatic water",
    "weather":    "weather thunder storm rain lightning atmospheric nature field-recording",
    "bell":       "bell marimba tubular-bell metallic resonant percussion melodic chime",
    "bass":       "bass sub-bass 808 low-frequency deep kick punch electronic",
    "glitch":     "glitch digital artifact bit-crush error noise stutter electronic experimental",
    "pinball":    "pinball arcade flipper bumper mechanical electromechanical game retro",
    "horror":     "horror dark-ambient tension scary dissonant drone suspense eerie atmospheric",
    "granular":   "granular grain cloud texture noise synthesizer generative abstract",
    "lofi":       "lofi lo-fi vinyl crackle cassette tape hiss analog warm degraded",
    "modem":      "modem dial-up internet dtmf handshake telecom data 90s digital",
    "insects":    "insects cricket cicada nature bug organic field-recording ambient summer",
    "gamelan":    "gamelan balinese javanese metalophone percussion world-music ethnic bell",
    "glass":      "glass crystal shard brittle resonant impact singing-rim sound-design",
    "clockwork":  "clockwork mechanism gear tick ratchet spring mechanical miniature",
    "creature":   "creature animal call chirp growl breath imaginary organic vocal",
    "cat":        "cat feline meow mewl purr hiss trill animal pet vocal procedural",
    "birds":      "birds birdsong peep chirp tweet whistle trill warble nature animal procedural",
    "electricity":"electricity electric arc transformer hum relay charge spark voltage",
    "cave":       "cave cavern drip stone rumble subterranean wind ambience reverb",
    "arp":        "arpeggio synthesizer melodic chord electronic music generative tonal",
}

BASE_TAGS = "glorb generative procedural python synthesizer sound-design"

GLORB_INFO = (
    "Generated by GLORB (https://bassimatte.github.io/glorb/), "
    "Matteo Bassi's browser-based generative audio synthesiser."
)

FS_LICENSE = "Creative Commons 0"
FS_PACK = "Glorb Generative Sounds by Bassimat"
FS_BST_CATEGORY = "fx-el"   # Sound effects > Electronic / Design on Freesound


# ── Rendering ────────────────────────────────────────────────────────────────

def render_mode(mode, duration, quality):
    """Render a mode and return a numpy audio array."""
    sr, subtype, _ = bb.QUALITY_PRESETS[quality]
    bb.SAMPLE_RATE = sr

    if mode == "glorb":
        parts, total = [], 0.0
        gap_lo = bb._knob(bb.KNOB_ENERGY, 0.22, 0.003)
        gap_hi = bb._knob(bb.KNOB_ENERGY, 0.80, 0.06)
        while total < duration:
            sig, _, _ = bb.make_blip()
            gap = random.uniform(gap_lo, gap_hi)
            parts.append(sig)
            parts.append(bb.silence(gap))
            total += len(sig) / bb.SAMPLE_RATE + gap
        return bb._finalise(np.concatenate(parts), duration), sr, subtype

    fn_map = {
        "retro":      bb.make_retro_sequence,
        "nature":     bb.make_nature_sequence,
        "scifi":      bb.make_scifi_sequence,
        "haptic":     bb.make_haptic_sequence,
        "radio":      bb.make_radio_sequence,
        "foley":      bb.make_foley_sequence,
        "underwater": bb.make_underwater_sequence,
        "weather":    bb.make_weather_sequence,
        "bell":       bb.make_bell_sequence,
        "bass":       bb.make_bass_sequence,
        "glitch":     bb.make_glitch_sequence,
        "pinball":    bb.make_pinball_sequence,
        "horror":     bb.make_horror_sequence,
        "granular":   bb.make_granular_sequence,
        "lofi":       bb.make_lofi_sequence,
        "modem":      bb.make_modem_sequence,
        "insects":    bb.make_insects_sequence,
        "gamelan":    bb.make_gamelan_sequence,
        "glass":      bb.make_glass_sequence,
        "clockwork":  bb.make_clockwork_sequence,
        "creature":   bb.make_creature_sequence,
        "cat":        bb.make_cat_sequence,
        "birds":      bb.make_birds_sequence,
        "electricity":bb.make_electricity_sequence,
        "cave":       bb.make_cave_sequence,
        "arp":        bb.make_arp_sequence,
    }

    if mode == "ui-pack":
        sounds = bb.make_ui_pack()
        parts = []
        for audio in sounds.values():
            peak = np.max(np.abs(audio))
            if peak > 0:
                audio = audio / peak * 0.9886
            parts.append(audio)
            parts.append(bb.silence(0.25))
        return np.concatenate(parts[:-1]), sr, subtype

    if mode == "nature":
        audio, _ = bb.make_nature_sequence(duration)
        return audio, sr, subtype

    fn = fn_map.get(mode)
    if fn is None:
        raise ValueError(f"Unknown mode: {mode}")
    return fn(duration), sr, subtype


def save_wav(audio, sample_rate, subtype, path):
    peak = np.max(np.abs(audio))
    if peak > 0:
        audio = audio / peak * 0.9886
    sf.write(str(path), audio.astype(np.float32), sample_rate, subtype=subtype)


# ── Freesound metadata ────────────────────────────────────────────────────────

def build_tags(mode):
    specific = MODE_TAGS.get(mode, "")
    combined = f"{BASE_TAGS} {specific}".split()
    seen, out = set(), []
    for t in combined:
        if t not in seen:
            seen.add(t); out.append(t)
    return " ".join(out[:30])


def build_description(mode):
    label, short_desc = MODES[mode]
    return f"{short_desc}.\n\n{GLORB_INFO}"


def build_title(mode):
    _label, short_desc = MODES[mode]
    return f"{short_desc} by GLORB"


def write_freesound_xlsx(rows, xlsx_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Freesound Bulk Describe"

    headers = ["audio_filename", "name", "tags", "geotag",
               "description", "license", "pack_name", "is_explicit", "bst_category"]

    # Header styling
    header_fill = PatternFill("solid", start_color="1F3864")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row[h] for h in headers])

    # Column widths
    widths = [35, 50, 80, 15, 80, 25, 35, 12, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # Wrap text for description column
    for row in ws.iter_rows(min_row=2):
        row[4].alignment = Alignment(wrap_text=True)

    wb.save(str(xlsx_path))


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Glorb — batch audio renderer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"Available modes: {', '.join(MODES.keys())}"
    )
    parser.add_argument("--mode", nargs="+", metavar="MODE",
                        help="Mode(s) to render")
    parser.add_argument("--all", action="store_true",
                        help="Render all modes")
    parser.add_argument("--duration", type=float, default=30.0, metavar="SECS",
                        help="Duration in seconds (default: 30)")
    parser.add_argument("--quality", choices=["standard", "high", "studio", "float"],
                        default="high", help="Audio quality preset (default: high)")
    parser.add_argument("--energy", type=int, default=50, metavar="0-100")
    parser.add_argument("--brightness", type=int, default=50, metavar="0-100")
    parser.add_argument("--chaos", type=int, default=50, metavar="0-100")
    parser.add_argument("--output-dir", default="exports/glorb", metavar="DIR")
    parser.add_argument("--freesound", action="store_true",
                        help="Generate Freesound bulk XLSX")
    parser.add_argument("--metadata-only", action="store_true",
                        help="Regenerate XLSX only (skip rendering)")
    parser.add_argument("--seed", type=int, default=None,
                        help="Random seed for reproducible renders")
    parser.add_argument("--list", action="store_true",
                        help="List available playback modes and sound presets, then exit")

    args = parser.parse_args()

    if args.list:
        bb.print_available_modes_and_presets()
        return

    # ── Determine modes to render ─────────────────────────────────────
    if args.all:
        selected = list(MODES.keys())
    elif args.mode:
        selected = []
        for m in args.mode:
            if m not in MODES:
                print(f"ERROR: Unknown mode '{m}'. Available: {', '.join(MODES.keys())}")
                sys.exit(1)
            selected.append(m)
    else:
        # Interactive picker
        print("\n── Glorb Modes ──────────────────────────────────────────────")
        for i, (k, (label, desc)) in enumerate(MODES.items(), 1):
            print(f"  {i:2}. {label:<12} {desc}")
        print("   0. All modes")
        print()
        raw = input("Enter mode number(s) separated by spaces (e.g. 1 3 5), or 0 for all: ").strip()
        if raw == "0":
            selected = list(MODES.keys())
        else:
            keys = list(MODES.keys())
            try:
                indices = [int(x) - 1 for x in raw.split()]
                selected = [keys[i] for i in indices]
            except (ValueError, IndexError):
                print("Invalid selection."); sys.exit(1)

    # ── Set knobs & seed ──────────────────────────────────────────────
    bb.KNOB_ENERGY     = max(0, min(100, args.energy))
    bb.KNOB_BRIGHTNESS = max(0, min(100, args.brightness))
    bb.KNOB_CHAOS      = max(0, min(100, args.chaos))

    seed = args.seed if args.seed is not None else random.randint(0, 999999)
    random.seed(seed)
    np.random.seed(seed)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    sr, subtype, quality_label = bb.QUALITY_PRESETS[args.quality]

    print(f"\n── Glorb batch render ───────────────────────────────────────")
    print(f"  Modes:      {', '.join(selected)}")
    print(f"  Duration:   {args.duration}s")
    print(f"  Quality:    {quality_label}")
    print(f"  Knobs:      Energy={bb.KNOB_ENERGY}  Brightness={bb.KNOB_BRIGHTNESS}  Chaos={bb.KNOB_CHAOS}")
    print(f"  Seed:       {seed}")
    print(f"  Output:     {output_dir.resolve()}")
    if args.metadata_only:
        print("  Mode:       --metadata-only (XLSX only, skip rendering)")
    print()

    rows = []
    total = len(selected)

    for i, mode in enumerate(selected, 1):
        label, desc = MODES[mode]
        safe = mode.replace("-", "_").replace(" ", "_")
        wav_name = f"Glorb_{safe}.wav"
        wav_path = output_dir / wav_name

        print(f"[{i:2}/{total}] {label}...", end="  ", flush=True)

        if not args.metadata_only:
            try:
                audio, sr_out, subtype_out = render_mode(mode, args.duration, args.quality)
                save_wav(audio, sr_out, subtype_out, wav_path)
                dur_actual = len(audio) / sr_out
                print(f"✓  ({dur_actual:.1f}s → {wav_name})")
            except Exception as e:
                print(f"ERROR: {e}")
                continue
        else:
            print("✓  (skipped)")

        rows.append({
            "audio_filename": wav_name,
            "name": build_title(mode),
            "tags": build_tags(mode),
            "geotag": "",
            "description": build_description(mode),
            "license": FS_LICENSE,
            "pack_name": FS_PACK,
            "is_explicit": "0",
            "bst_category": FS_BST_CATEGORY,
        })

    print()

    if args.freesound or args.metadata_only:
        xlsx_path = output_dir / "freesound_bulk.xlsx"
        print(f"Writing Freesound XLSX → {xlsx_path}")
        write_freesound_xlsx(rows, xlsx_path)
        print(f"✓  {len(rows)} rows written.")
        print(f"\nUpload audio files, then use XLSX at: https://freesound.org/home/describe/")

    print(f"\n✓ Done! {len(rows)} file(s) in {output_dir.resolve()}")


if __name__ == "__main__":
    main()
